# -*- coding: utf-8 -*-
"""
Created on Tue Nov 22 16:47:35 2016

@author: nicol
"""

from pulp import *
import openpyxl
from openpyxl import *

wb = openpyxl.load_workbook('data.xlsx')

def CaseDisaster():
    
    SheetTracts = wb['Tract]s']
    SheetHospitals = wb['Hospitals']
    SheetDistances = wb['Distances']
    SheetExample = wb['Example']
    SheetCaseOne = wb['Case 1']
    
    di = [i for i in range (1,824)]
    dj = [j for j in range (1,57)]


    # Create the 'prob' variable to contain the problem data
#    prob = pulp.LpProblem("Disaster Case",LpMinimize)
    
    # Create the inputs for the problem data

    Z = list(SheetHospitals.cell(row=j, column=3).value for j in range(1,58))


   
#   Create Dij
    D = [[0]]
    for i in range (0,823) :
        D.append([])
        for j in range (56*i+1,56*(i+1)+2) :
            D[i+1].append(SheetDistances.cell(row=j+1, column=9).value)




# Number of people that needed help according to Hazus
    F = list(SheetExample.cell(row=j+1, column=9).value for j in range(0,826))

    prob = LpProblem("Disaster Case",LpMinimize)
    
# Creation of the DV
    choices = LpVariable.dicts("X",(di,dj),0,14000,LpInteger)

            
# Obj fct
    prob += lpSum(D[i][j]*choices[i][j]for i in di for j in dj )

# constraint 1
    for i in di:
        prob += lpSum(choices[i][j]  for j in dj) <= F[i]
    
# constraint 2
    for j in dj:
        prob += lpSum(choices[i][j]  for i in di) == Z[j]
    
    
   # The problem data is written to an .lp file
    prob.writeLP("Disaster Case.lp")

# The problem is solved using PuLP's choice of Solver
    prob.solve()

# The status of the solution is printed to the screen
    print ("Status:", LpStatus[prob.status])

# Each of the variables is printed with it's resolved optimum value
    for v in prob.variables():
        print (v.name, "=", v.varValue)


# The optimised objective function value is printed to the screen    
    print ("Optimal Solution ", value(prob.objective))
    
    for i in range(0,823) :
        for j in range(0,56) :
            SheetDistances.cell(row=3+56*i+j,column=11).value=str(X[i][j])
            SheetDistances.cell(row=3+56*i+j,column=12).value=X[i][j].varValue
    wb.save('solution.xlsx')

        
caseDisaster()
